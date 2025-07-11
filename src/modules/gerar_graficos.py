import sys
from pathlib import Path
from io import BytesIO
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import numpy as np
import math
import seaborn as sns
from datetime import datetime

# Garante que o diretório 'src' esteja no PYTHONPATH
sys.path.append(str(Path(__file__).resolve().parents[1]))
from src.config.settings import CAMINHO_PLANILHA_FINAL

# Mapeamento de nomes completos para siglas
meses_sigla = {
    "Janeiro": "Jan", "Fevereiro": "Fev", "Março": "Mar", "Abril": "Abr",
    "Maio": "Mai", "Junho": "Jun", "Julho": "Jul", "Agosto": "Ago",
    "Setembro": "Set", "Outubro": "Out", "Novembro": "Nov", "Dezembro": "Dez"
}
ordem_meses = list(meses_sigla.keys())
siglas_meses = list(meses_sigla.values())
goves_nome = "GOVES - ESTADO DO ESPIRITO SANTO"

IMG_WIDTH_LINHA = 580
IMG_HEIGHT_LINHA = 420
IMG_WIDTH_BARRAS = 880
IMG_HEIGHT_BARRAS = 520
IMG_WIDTH_HEATMAP = 980
IMG_HEIGHT_HEATMAP = 820

mensagens = []

def gerar_dicionario_fluxos(arquivo_excel):
    resultado = {}

    planilhas = pd.read_excel(arquivo_excel, sheet_name=None, skiprows=5)

    for nome_aba, df in planilhas.items():
        nome_mes = nome_aba.strip().capitalize()
        if nome_mes not in ordem_meses:
            continue

        if not {'Patriarca', 'Quantidade'}.issubset(df.columns):
            continue

        df = df.dropna(subset=['Patriarca', 'Quantidade'])
        agrupado = df.groupby('Patriarca')['Quantidade'].sum()
        resultado[nome_mes] = agrupado.to_dict()

    return resultado

def carregar_dados_setores_goves(caminho_arquivo):
    planilhas = pd.read_excel(caminho_arquivo, sheet_name=None, skiprows=5)
    df_completo = []

    for nome_aba, df in planilhas.items():
        mes = nome_aba.strip().capitalize()
        if mes not in ordem_meses or 'Patriarca' not in df.columns or 'Setor' not in df.columns or 'Quantidade' not in df.columns:
            continue

        df = df.dropna(subset=['Patriarca', 'Setor', 'Quantidade'])
        df_goves = df[df['Patriarca'] == goves_nome].copy()
        df_goves['Mês'] = meses_sigla[mes]
        df_completo.append(df_goves[['Setor', 'Quantidade', 'Mês']])

    if df_completo:
        return pd.concat(df_completo)
    else:
        return pd.DataFrame(columns=['Setor', 'Quantidade', 'Mês'])

def gerar_grafico_linha_fluxos_mensais(caminho_arquivo: str):
    fluxos_publicados = []
    excel_file = pd.ExcelFile(caminho_arquivo)
    abas = excel_file.sheet_names

    for mes in ordem_meses:
        if mes in abas:
            df = pd.read_excel(caminho_arquivo, sheet_name=mes, skiprows=6, header=None)
            quantidade = pd.to_numeric(df.iloc[:, 2], errors="coerce").dropna()
            total = int(quantidade.sum())
        else:
            total = 0
        fluxos_publicados.append(total)

    fig, ax = plt.subplots(figsize=(10, 5))
    plt.plot(siglas_meses, fluxos_publicados, marker="o", linestyle="-", color="blue")
    plt.xlabel("Meses")
    plt.ylabel("Total de Fluxos Publicados")
    plt.title(f"Evolução Mensal de Fluxos Publicados em {datetime.now().year}")
    plt.grid(True)
    plt.yticks(range(0, max(fluxos_publicados) + 50, 50))
    plt.tight_layout()

    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=150)
    plt.close(fig)
    buffer.seek(0)

    wb = load_workbook(caminho_arquivo)
    if "Graficos" in wb.sheetnames:
        del wb["Graficos"]
    aba_graficos = wb.create_sheet("Graficos")
    img = XLImage(buffer)
    img.width = IMG_WIDTH_LINHA
    img.height = IMG_HEIGHT_LINHA
    aba_graficos.add_image(img, "A1")
    wb.save(caminho_arquivo)

    return {"status": "Sucesso Gráfico de Linha", "mensagem": "Gráfico de linha salvo com sucesso na aba 'Graficos'."}

def gerar_grafico_barras(dados_por_mes):
    dados_completos = {mes: dados_por_mes.get(mes, {}) for mes in ordem_meses}
    siglas_meses = [meses_sigla[mes] for mes in ordem_meses]

    # Coleta e ordena patriarcas, com GOVES no início
    patriarcas = sorted({p for dados in dados_completos.values() for p in dados})
    if goves_nome in patriarcas:
        patriarcas.remove(goves_nome)
        patriarcas = [goves_nome] + patriarcas

    n_patriarcas = len(patriarcas)

    valores_por_patriarca = {p: [] for p in patriarcas}
    for mes in ordem_meses:
        dados_mes = dados_completos[mes]
        for p in patriarcas:
            valores_por_patriarca[p].append(dados_mes.get(p, 0))

    x = np.arange(len(ordem_meses))

    # Largura das barras proporcional ao número de patriarcas
    bar_width = min(0.9 / max(n_patriarcas, 1), 0.3)
    deslocamento_total = bar_width * n_patriarcas

    fig, ax = plt.subplots(figsize=(14, 6))

    for i, (patriarca, valores) in enumerate(valores_por_patriarca.items()):
        deslocamento = i * bar_width
        barras = ax.bar(x + deslocamento, valores, width=bar_width, label=patriarca)

        # Coloca valores acima das barras
        for xi, yi in zip(x + deslocamento, valores):
            if yi > 0:
                ax.text(xi, yi + 0.5, str(int(yi)), ha='center', va='bottom', fontsize=8)

    ax.set_xticks(x + deslocamento_total / 2 - bar_width / 2)
    ax.set_xticklabels(siglas_meses)
    ax.set_xlabel("Mês")
    ax.set_ylabel("Total de Publicações")
    ax.set_title("Publicações por Patriarca ao Longo do Ano")

    ncolunas_legenda = math.ceil(n_patriarcas / 4)

    ax.legend(
        title="Patriarca",
        loc="upper center",
        bbox_to_anchor=(0.5, -0.15),
        ncol=ncolunas_legenda,
        fontsize=9,
        title_fontsize=10
    )

    ax.grid(axis='y', linestyle='--', alpha=0.6)
    plt.tight_layout()

    # Salva a figura em um buffer de memória
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=150)
    plt.close(fig)
    buffer.seek(0)

    # Insere na aba "Graficos" da planilha
    wb = load_workbook(CAMINHO_PLANILHA_FINAL)
    aba_graficos = wb["Graficos"] if "Graficos" in wb.sheetnames else wb.create_sheet("Graficos")

    # Insere a imagem no canto direito (coluna J)
    img = XLImage(buffer)
    img.width = IMG_WIDTH_BARRAS
    img.height = IMG_HEIGHT_BARRAS
    aba_graficos.add_image(img, "J1")

    wb.save(CAMINHO_PLANILHA_FINAL)

    return {"status": "Sucesso Gráfico de Barras", "mensagem": "Gráfico de barras salvo com sucesso na aba 'Graficos'."}

def gerar_heatmap_setores_por_mes(caminho_arquivo: str):
    df = carregar_dados_setores_goves(caminho_arquivo)
    if df.empty:
        return {"status": "Falha ao gerar o gráfico HeatMap",
                "mensagem": f"Não foi possivel gerar o gráfico HeatMap, nenhum dado do GOVES encontrado."}

    tabela = df.pivot_table(index='Setor', columns='Mês', values='Quantidade', aggfunc='sum', fill_value=0)
    tabela = tabela[[sigla for sigla in siglas_meses if sigla in tabela.columns]]

    fig, ax = plt.subplots(figsize=(14, max(6, len(tabela) * 0.4)))
    sns.heatmap(tabela, annot=True, fmt=".0f", cmap="Blues", linewidths=0.5, cbar_kws={'label': 'Publicações'}, ax=ax)

    azul_escuro = '#003366'
    ax.set_title("Heatmap - Publicações por Setor (GOVES) ao Longo dos Meses", fontsize=12, fontweight='bold', color=azul_escuro)
    ax.set_xlabel("Meses", fontsize=12, fontweight='bold', color=azul_escuro)
    ax.set_ylabel("Setores", fontsize=12, fontweight='bold', color=azul_escuro)
    cbar = ax.collections[0].colorbar
    cbar.set_label("Publicações", fontsize=12, fontweight='bold', color=azul_escuro)
    plt.tight_layout()

    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=150)
    plt.close(fig)
    buffer.seek(0)

    wb = load_workbook(caminho_arquivo)
    aba_graficos = wb["Graficos"] if "Graficos" in wb.sheetnames else wb.create_sheet("Graficos")

    img = XLImage(buffer)
    img.width = IMG_WIDTH_HEATMAP
    img.height = IMG_HEIGHT_HEATMAP
    aba_graficos.add_image(img, "A28")

    wb.save(caminho_arquivo)
    return {"status": f"Sucesso Grafico HeatMap gerado com sucesso",
            "mensagem": f"Heatmap salvo com sucesso na aba 'Graficos'."}

def gerar_graficos_gerais():
    mensagens = []

    resultado = gerar_grafico_linha_fluxos_mensais(CAMINHO_PLANILHA_FINAL)
    mensagens.append(resultado)

    resultado = gerar_grafico_barras(gerar_dicionario_fluxos(CAMINHO_PLANILHA_FINAL))
    mensagens.append(resultado)

    resultado = gerar_heatmap_setores_por_mes(CAMINHO_PLANILHA_FINAL)
    mensagens.append(resultado)

    status = resultado["status"]
    if "falha" in status.lower():
       return mensagens

    mensagens.append({
        "status": "Sucesso na Geração dos Gráficos", "mensagem": "Todos os gráficos foram gerados com sucesso na aba 'Gráficos'."})

    return mensagens

