import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment

# Garante que o diretório 'src' esteja no PYTHONPATH
sys.path.append(str(Path(__file__).resolve().parents[1]))

from src.config.settings import CAMINHO_PLANILHA_FINAL, ABA_MODELO
from src.utils.excel_helpers import ajustar_largura_colunas


def padronizar_coluna_nome(mes: str, df: pd.DataFrame) -> pd.DataFrame:
    if "Nome" in df.columns:
        return df

    if "Nomes" in df.columns:
        return df.rename(columns={"Nomes": "Nome"})

    return {"status": f"Falha no processamento do mes {mes}", "mensagem": f"A planilha com as informações do mes de {mes} deve conter a coluna 'Nome' ou 'Nomes'."}


def limpar_dados_antigos(aba: Worksheet, linha_inicial: int = 7, col_final: int = 5):
    for row in aba.iter_rows(min_row=linha_inicial, max_row=aba.max_row, min_col=1, max_col=col_final):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)


def gerar_fluxo_mensal(mes: str, caminho_fluxos: str):
    caminho_fluxos = Path(caminho_fluxos)

    # === LEITURA DA PLANILHA DE FLUXOS ===
    df = pd.read_excel(caminho_fluxos, engine="openpyxl")
    df = padronizar_coluna_nome(mes, df)
    if isinstance(df, dict):  # Se retornou erro
        return df

    colunas_necessarias = {"Patriarca", "Órgão", "Nome"}
    if not colunas_necessarias.issubset(set(df.columns)):
        return {"status": f"Falha no processamento do mes {mes}",
                "mensagem": f"A planilha deve conter as colunas: {', '.join(colunas_necessarias)}."}

    df = df.dropna(subset=colunas_necessarias)

    df_resumo = (
        df.groupby(["Patriarca", "Órgão"])["Nome"]
        .count()
        .reset_index()
        .rename(columns={"Órgão": "Setor", "Nome": "Quantidade"})
        .sort_values(by=["Patriarca", "Setor"])
    )

    # === ABRIR A PLANILHA DE DESTINO ===
    wb = load_workbook(CAMINHO_PLANILHA_FINAL)

    if ABA_MODELO not in wb.sheetnames:
        return {"status": f"Falha no processamento do mes {mes}",
                "mensagem": f"A planilha deve conter uma aba chamada '{ABA_MODELO}' como modelo."}

    if mes in wb.sheetnames:
        del wb[mes]

    aba_modelo: Worksheet = wb[ABA_MODELO]
    nova_aba: Worksheet = wb.copy_worksheet(aba_modelo)
    nova_aba.title = mes

    # Inserir nova aba antes da aba "Graficos", se existir
    if "Graficos" in wb.sheetnames:
        idx_graficos = wb.sheetnames.index("Graficos")
        wb.remove(nova_aba)
        wb._add_sheet(nova_aba, index=idx_graficos)  # Ainda é protegido, mas mais seguro que _sheets

    # Atualiza cabeçalhos
    nova_aba["A4"] = f"Contagem de Fluxos Publicados e Ativos por Setor - Mês {mes}"
    nova_aba["A5"] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    nova_aba["A6"] = "Patriarca"
    nova_aba["B6"] = "Setor"
    nova_aba["C6"] = "Quantidade"
    nova_aba["E6"] = "Total"

    # Remove dados anteriores
    limpar_dados_antigos(nova_aba)

    # Estilos base
    fill_branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_azul = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    alinhamento_esquerda = Alignment(horizontal="left")

    # Preenche dados a partir da linha 7
    linha = 7
    for i, (_, row) in enumerate(df_resumo.iterrows()):
        cor = fill_branco if i % 2 == 0 else fill_azul
        nova_aba[f"A{linha}"] = row["Patriarca"]
        nova_aba[f"B{linha}"] = row["Setor"]
        nova_aba[f"C{linha}"] = row["Quantidade"]

        for col in ["A", "B", "C"]:
            cell = nova_aba[f"{col}{linha}"]
            cell.fill = cor
            if col == "B":
                cell.alignment = alinhamento_esquerda

        linha += 1

    # Soma total na coluna E
    nova_aba["E7"] = f"=SUM(C7:C{linha-1})"

    # Ajusta largura das colunas
    ajustar_largura_colunas(nova_aba, colunas=["A", "B", "C"])

    # Salva
    wb.save(CAMINHO_PLANILHA_FINAL)
    return {"status": f"Sucesso no processamento do mes {mes}",
            "mensagem": f"Aba '{mes}' criada/atualizada com sucesso no arquivo:\n{CAMINHO_PLANILHA_FINAL}."}
